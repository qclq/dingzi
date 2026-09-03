import asyncio
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.core.config import Settings, get_settings
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models.detection import Defect, Detection
from app.models.export import ExportJob
from app.models.user import User
from app.security.auth import create_token, hash_password
from app.services import exports as export_service


@pytest.fixture()
def history_client():
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    sessions = async_sessionmaker(engine, expire_on_commit=False)

    async def setup() -> None:
        async with engine.begin() as connection:
            await connection.run_sync(Base.metadata.create_all)
        async with sessions() as session:
            session.add(
                User(
                    username="history-user",
                    password_hash=hash_password("History123!"),
                    display_name="历史测试",
                    role="operator",
                )
            )
            start = datetime(2026, 1, 1, tzinfo=UTC)
            for index in range(45):
                detection = Detection(
                    image_id=f"history-{index:03}",
                    line_id="line-1",
                    captured_at=start + timedelta(minutes=index),
                    operator="operator-a",
                    result="NG" if index % 2 else "PASS",
                    image_path=f"/data/images/archive/history-{index:03}.png",
                    thumbnail_path=None,
                    model_version="model-test",
                    config_version="config-test",
                    inference_ms=12.5,
                    defect_count=1,
                    raw_output={"index": index},
                )
                detection.defects = [
                    Defect(
                        type="scratch",
                        level="severe" if index % 2 else "minor",
                        confidence=0.91,
                        bbox=[0.1, 0.2, 0.3, 0.4],
                        width_mm=1.2,
                        height_mm=0.8,
                    )
                ]
                session.add(detection)
            await session.commit()

    asyncio.run(setup())

    async def override_db():
        async with sessions() as session:
            yield session

    app.dependency_overrides[get_db] = override_db
    token, _ = create_token(get_settings(), 1, "operator", "access", timedelta(hours=1))
    with TestClient(app) as client:
        yield client, sessions, {"Authorization": f"Bearer {token}"}
    app.dependency_overrides.clear()
    asyncio.run(engine.dispose())


def test_detection_pagination_uses_descending_order(history_client) -> None:
    client, _, headers = history_client
    response = client.get("/api/v1/detections?page=2&page_size=20", headers=headers)

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 45
    assert len(payload["items"]) == 20
    assert payload["items"][0]["image_id"] == "history-024"
    assert payload["items"][-1]["image_id"] == "history-005"


def test_detection_filters(history_client) -> None:
    client, _, headers = history_client
    response = client.get(
        "/api/v1/detections",
        headers=headers,
        params={
            "start_time": "2026-01-01T00:10:00Z",
            "end_time": "2026-01-01T00:20:00Z",
            "result": "NG",
            "page_size": 20,
        },
    )

    assert response.status_code == 200
    assert [item["image_id"] for item in response.json()["items"]] == [
        "history-019",
        "history-017",
        "history-015",
        "history-013",
        "history-011",
    ]
    exact = client.get(
        "/api/v1/detections",
        headers=headers,
        params={"operator": "operator-a", "image_id": "history-010", "line_id": "line-1"},
    )
    assert exact.status_code == 200
    assert [item["image_id"] for item in exact.json()["items"]] == ["history-010"]


def test_detection_detail_and_signed_file_url(history_client) -> None:
    client, _, headers = history_client
    detail = client.get("/api/v1/detections/1", headers=headers)
    signed = client.get("/api/v1/detections/1/files/image", headers=headers)

    assert detail.status_code == 200
    assert detail.json()["raw_output"] == {"index": 0}
    assert detail.json()["defects"][0]["width_mm"] == 1.2
    assert signed.status_code == 200
    assert signed.json()["url"].startswith("http://localhost:9000/dingzi-files/")
    assert "expires=" in signed.json()["url"]
    assert "signature=" in signed.json()["url"]


def test_link_mes_work_order(history_client) -> None:
    client, _, headers = history_client

    linked = client.patch(
        "/api/v1/detections/1/mes-work-order",
        headers=headers,
        json={"mes_work_order": "MES-WO-2026-001"},
    )

    assert linked.status_code == 200
    assert linked.json()["mes_work_order"] == "MES-WO-2026-001"
    assert client.get("/api/v1/detections/1", headers=headers).json()["mes_work_order"] == "MES-WO-2026-001"


def test_detection_validation_and_authentication(history_client) -> None:
    client, _, headers = history_client

    assert client.get("/api/v1/detections").status_code == 401
    assert client.get("/api/v1/detections?page_size=10", headers=headers).status_code == 422
    assert client.get(
        "/api/v1/detections",
        headers=headers,
        params={"start_time": "2026-02-01T00:00:00Z", "end_time": "2026-01-01T00:00:00Z"},
    ).status_code == 422
    assert client.get("/api/v1/detections/99999", headers=headers).status_code == 404


def test_export_api_returns_without_running_report(history_client, monkeypatch: pytest.MonkeyPatch) -> None:
    client, _, headers = history_client
    dispatched: list[str] = []
    monkeypatch.setattr(
        "app.api.v1.exports.generate_detection_export.delay",
        lambda export_id: dispatched.append(export_id),
    )

    created = client.post(
        "/api/v1/exports",
        headers=headers,
        json={"format": "xlsx", "result": "NG"},
    )

    assert created.status_code == 202
    assert created.json()["status"] == "queued"
    assert dispatched == [created.json()["id"]]
    status_response = client.get(f"/api/v1/exports/{created.json()['id']}", headers=headers)
    assert status_response.status_code == 200
    assert status_response.json()["status"] == "queued"


@pytest.mark.parametrize("export_format", ["xlsx", "pdf"])
def test_generate_excel_and_pdf_exports(
    history_client,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    export_format: str,
) -> None:
    _, sessions, _ = history_client
    settings = Settings(
        jwt_secret_key="0123456789abcdef0123456789abcdef",
        export_directory=str(tmp_path),
    )
    monkeypatch.setattr(export_service, "get_settings", lambda: settings)
    monkeypatch.setattr(export_service, "EXPORT_BATCH_SIZE", 10)

    async def generate() -> tuple[ExportJob, Path]:
        async with sessions() as session:
            job = ExportJob(
                id=f"export-{export_format}",
                created_by_id=1,
                format=export_format,
                status="queued",
                query={"result": "NG"},
            )
            session.add(job)
            await session.commit()
            await export_service.generate_export(session, job.id)
            await session.refresh(job)
            return job, Path(job.file_path or "")

    job, output = asyncio.run(generate())

    assert job.status == "completed"
    assert job.record_count == 22
    assert output.exists() and output.stat().st_size > 100
    if export_format == "xlsx":
        workbook = load_workbook(output, read_only=True)
        assert sum(1 for _ in workbook["检测记录"].iter_rows()) == 23
    else:
        assert output.read_bytes().startswith(b"%PDF")
